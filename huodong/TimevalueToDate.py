#!usr/bin/env python
# -*- coding:utf-8 -*-

from openpyxl import load_workbook
import datetime,time

excel = load_workbook(r'E:\呱呱财经\测试报告\元宝消费原始数据.xlsx')

table = excel.get_sheet_by_name('Sheet5')

nrows = table.max_row
ncols = table.max_column

print (nrows,ncols)
# print (type(time.time()))
# testnum = 1500262276683
# time.localtime(1500262276683)

for i in range(2,nrows+1):
    timevalue = int(table.cell(row=i,column=2).value/1000)
    dateArray = time.localtime(timevalue)
    datestr = time.strftime("%Y/%m/%d",dateArray)
    # print (type(timevalue))
    table.cell(row=i,column=1).value=datestr
    print (timevalue,datestr)
excel.save(r'E:\呱呱财经\测试报告\元宝消费原始数据.xlsx')
